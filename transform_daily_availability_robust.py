import argparse
import logging
import re
import sys
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from operations_kpi_logging import add_log_level_arg, configure_logging

logger = logging.getLogger("operations_kpi.etl.transform_daily_availability_robust")


def _read_excel_raw(file_path, sheet_name=None, data_only=False):
    """Read worksheet values into a DataFrame, optionally using cached formula results."""
    if not data_only:
        if sheet_name:
            return pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        return pd.read_excel(file_path, header=None)

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    return pd.DataFrame(rows)


def transform_daily_availability(
    file_path,
    sheet_name=None,
    fixed_cols_end=None,
    header_row=2,
    data_start_row=3,
    data_only=False,
):
    """
    Transform Excel file with merged date headers to relational format.
    
    Parameters:
    -----------
    file_path : str
        Path to the Excel file
    sheet_name : str, optional
        Name of the sheet/tab to read. If None, reads the first sheet.
    fixed_cols_end : int, optional
        Column index where date columns start (0-indexed). If None, will auto-detect.
    header_row : int, optional
        Row index containing fixed column headers (default: 2)
    data_start_row : int, optional
        Row index where data starts (default: 3)
    data_only : bool, optional
        If True, read cached formula results from the workbook instead of formulas.
    
    Returns:
    --------
    pd.DataFrame
        Transformed data in long format
    """
    
    # Read Excel file
    df_raw = _read_excel_raw(file_path, sheet_name=sheet_name, data_only=data_only)
    
    # Auto-detect fixed columns end if not provided
    if fixed_cols_end is None:
        # Look for "Availability" column in header row (row 2)
        if header_row < len(df_raw):
            for col_idx in range(len(df_raw.columns)):
                cell_value = str(df_raw.iloc[header_row, col_idx]).lower()
                if 'availability' in cell_value:
                    fixed_cols_end = col_idx + 1
                    break
        
        # Fallback: look for datetime objects in row 0 (date headers)
        if fixed_cols_end is None:
            for col_idx in range(len(df_raw.columns)):
                cell_value = df_raw.iloc[0, col_idx]
                if isinstance(cell_value, (datetime, pd.Timestamp)) or (pd.notna(cell_value) and 'datetime' in str(type(cell_value))):
                    fixed_cols_end = col_idx
                    break
        
        if fixed_cols_end is None:
            raise ValueError("Could not auto-detect where date columns start. Please specify fixed_cols_end parameter.")
    
    logger.debug("Fixed columns end at column index: %s", fixed_cols_end)
    
    # Extract fixed columns
    if data_only:
        header_values = list(df_raw.iloc[header_row, :fixed_cols_end].values)
        fixed_values = df_raw.iloc[data_start_row:, :fixed_cols_end].copy()
        fixed_df = pd.DataFrame(fixed_values.values, columns=header_values)
    else:
        if sheet_name:
            fixed_df = pd.read_excel(
                file_path, sheet_name=sheet_name, usecols=range(fixed_cols_end), header=header_row
            )
        else:
            fixed_df = pd.read_excel(file_path, usecols=range(fixed_cols_end), header=header_row)
    fixed_df = fixed_df.dropna(how='all')  # Remove completely empty rows
    
    logger.debug("Fixed columns: %s", list(fixed_df.columns))
    logger.debug("Number of fixed data rows: %d", len(fixed_df))
    
    # Parse date columns from row 0
    date_columns_info = []
    row0 = df_raw.iloc[0, fixed_cols_end:].values
    
    # Dates appear every 3 columns (merged cells)
    # Pattern: Date, NaN, NaN, Date, NaN, NaN, ...
    for i in range(0, len(row0), 3):
        if i + 2 >= len(row0):
            break
        
        # Date is in the first column of each group of 3
        date_cell = row0[i]
        
        # Check if it's a datetime object or can be parsed as date
        if pd.notna(date_cell):
            # If it's already a datetime, use it directly
            if isinstance(date_cell, (datetime, pd.Timestamp)):
                date_value = date_cell
            else:
                # Try to parse as datetime
                try:
                    date_value = pd.to_datetime(date_cell)
                except (ValueError, TypeError):
                    date_value = None
            
            if date_value is not None:
                date_columns_info.append({
                    'date': date_value,
                    'col_start': fixed_cols_end + i,
                })
    
    logger.debug("Found %d date columns", len(date_columns_info))
    
    # Transform data
    result_rows = []
    
    # Data starts at data_start_row
    num_data_rows = len(df_raw) - data_start_row
    
    logger.debug(
        "Processing %d data rows starting from row %d",
        num_data_rows,
        data_start_row,
    )
    
    for date_info in date_columns_info:
        date = date_info['date']
        col_start = date_info['col_start']
        
        for row_idx in range(num_data_rows):
            actual_row = data_start_row + row_idx
            
            if row_idx >= len(fixed_df):
                break
            
            # Skip if fixed row is completely empty
            if fixed_df.iloc[row_idx].isna().all():
                continue
            
            # Extract the three metric values
            incident_count = None
            outage_mins = None
            uptime_per_tenant = None
            
            if col_start < len(df_raw.columns):
                incident_count = df_raw.iloc[actual_row, col_start]
            if col_start + 1 < len(df_raw.columns):
                outage_mins = df_raw.iloc[actual_row, col_start + 1]
            if col_start + 2 < len(df_raw.columns):
                uptime_per_tenant = df_raw.iloc[actual_row, col_start + 2]
            
            # Create combined row
            row_dict = fixed_df.iloc[row_idx].to_dict()
            row_dict['Date'] = date
            row_dict['Incident_count'] = incident_count
            row_dict['Outage_mins'] = outage_mins
            row_dict['Uptime_per_tenant'] = uptime_per_tenant
            
            result_rows.append(row_dict)
    
    # Create result DataFrame
    result_df = pd.DataFrame(result_rows)
    
    # Clean up
    result_df = result_df.dropna(subset=['Date'])
    
    # Convert numeric columns
    if 'Incident_count' in result_df.columns:
        result_df['Incident_count'] = pd.to_numeric(result_df['Incident_count'], errors='coerce').fillna(0).astype(int)
    if 'Outage_mins' in result_df.columns:
        result_df['Outage_mins'] = pd.to_numeric(result_df['Outage_mins'], errors='coerce').fillna(0.0)
    
    # Sort by fixed columns and date
    sort_cols = [col for col in fixed_df.columns if col in result_df.columns] + ['Date']
    result_df = result_df.sort_values(by=sort_cols)
    
    # Reset index
    result_df = result_df.reset_index(drop=True)
    
    return result_df


def main() -> None:
    parser = argparse.ArgumentParser(description="Transform daily availability Excel to CSV.")
    parser.add_argument("--input", default="daily_availability_jan_26.xlsx")
    parser.add_argument("--sheet", default="Daily Site Availability")
    parser.add_argument("--output", default="daily_availability_jan_26_transformed.csv")
    add_log_level_arg(parser)
    args = parser.parse_args()
    configure_logging(args.log_level)

    input_file = args.input
    sheet_name = args.sheet
    output_file = args.output

    logger.info("Daily Availability Data Transformer")
    logger.info("Input file: %s", input_file)
    logger.info("Sheet name: %s", sheet_name)

    try:
        df_transformed = transform_daily_availability(input_file, sheet_name=sheet_name)

        logger.info(
            "Output shape: %d rows x %d columns",
            df_transformed.shape[0],
            df_transformed.shape[1],
        )
        logger.debug("Columns: %s", list(df_transformed.columns))
        logger.debug("First rows:\n%s", df_transformed.head(10).to_string())
        logger.info(
            "Date range: %s .. %s (%d unique dates)",
            df_transformed["Date"].min(),
            df_transformed["Date"].max(),
            df_transformed["Date"].nunique(),
        )

        logger.info("Saving to %s", output_file)
        df_transformed.to_csv(output_file, index=False)

        output_excel = output_file.replace(".csv", ".xlsx")
        df_transformed["Date"] = pd.to_datetime(df_transformed["Date"])
        df_transformed.to_excel(output_excel, index=False)

        logger.info("Success! Output saved to %s and %s", output_file, output_excel)

    except FileNotFoundError:
        logger.error("File not found: %s", input_file)
        raise SystemExit(1) from None
    except ValueError as e:
        if "Worksheet named" in str(e) or "sheet" in str(e).lower():
            logger.error("Sheet %r not found in %s", sheet_name, input_file)
            try:
                xl_file = pd.ExcelFile(input_file)
                logger.error("Available sheets: %s", ", ".join(xl_file.sheet_names))
            except Exception:
                logger.exception("Could not list workbook sheets")
            raise SystemExit(1) from e
        logger.exception("Transformation failed")
        raise SystemExit(1) from e
    except Exception:
        logger.exception("Transformation failed")
        raise SystemExit(1) from None


if __name__ == "__main__":
    main()
